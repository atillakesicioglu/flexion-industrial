# -*- coding: utf-8 -*-
"""
Excel'den tum urun verilerini + gorselleri cikart.
Cikti:
  scripts/output/products.json        - urun verileri
  scripts/output/images/rowN_B.png    - urun ana gorseli
  scripts/output/images/rowN_X_0.png  - ikon gorseli
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json, io, os, re, unicodedata

BASE = r'C:\Users\atill\OneDrive\MASAST~1\projects\flexion'
XLSX_PATH = os.path.join(BASE, 'RNLER~1.XLS')
OUT_DIR = os.path.join(BASE, 'scripts', 'output')
IMG_DIR = os.path.join(OUT_DIR, 'images')

os.makedirs(IMG_DIR, exist_ok=True)

# --- Yardimci: satirdan onemi var mi ---
HEADER_KEYWORDS = {
    'WATER HOSES', 'AIR-GAS HOSES', 'AIR - GAS HOSES', 'OIL & PETROLEUM HOSES',
    'WELDING HOSES', 'FOOD HOSES', 'FOOD & BEVERAGE HOSES', 'MATERIAL HANDLING HOSES',
    'MATERIAL HANDLING', 'SEWER CLEANING HOSES', 'SEWER & ROAD CLEANING',
    'SEWER JETTING & CLEANING HOSES', 'STEAM HOSES', 'CHEMICAL HOSES', 'HYDRAULIC HOSES',
    'HYDRAULIK HOSES', 'SILICONE HOSES', 'THERMOPLASTIC HOSES', 'PVC HOSES',
    'OFFSHORE HOSES', 'STEEL MILL HOSES', 'FIRE FIGHTING HOSES', 'CABLE PROTECTION HOSES',
    'HOT WATER HOSES', 'HOT AIR BLOWER HOSES', 'ABRASIVE FOOD HOSES',
    'PETROLEUM DISPENSING', 'FLEXION INDUSTRIAL', 'WATER HOSES',
    'AIR-GAS', 'CHEMICAL', 'FOOD', 'STEAM', 'WELDING', 'OIL & PETROLEUM',
    'HYDRAULIC', 'SILICONE', 'OFFSHORE', 'STEEL MILL', 'FIRE FIGHTING',
    'CABLE PROTECTION', 'HOT WATER', 'ABRASIVE FOOD', 'ALIZE 200\u00b0C',
    'ALIZE SD 200\u00b0C', 'PVC HOSES', 'THERMOPLASTIC HOSES',
}

def clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    # Normalize line breaks
    s = s.replace('\r\n', '\n').replace('\r', '\n')
    return s

def is_header_row(row_vals):
    """Baslik satirini veya bos satiri atla."""
    m = row_vals[12] if len(row_vals) > 12 else None  # M = Item Code (index 12, 0-based)
    n = row_vals[13] if len(row_vals) > 13 else None  # N = Name
    k = row_vals[10] if len(row_vals) > 10 else None  # K = Category
    # Eger M ve N bos, sadece K dolu -> baslik satiri
    if not m and not n:
        return True
    k_str = clean(k).upper()
    if k_str in HEADER_KEYWORDS:
        # Eger sadece K var ve diger alanlar bos -> baslik
        if not m and not n:
            return True
    return False

def build_html_description(row_vals):
    """O..W sutunlarindan HTML aciklama olustur."""
    # Sutun endeksleri (0-based)
    # O=14, P=15, Q=16, R=17, S=18, T=19, U=20, V=21, W=22
    sections = [
        (14, 'Description of Goods'),
        (15, 'WP [Bar] / Size [mm] [Inch]'),
        (16, None),   # Size - WP ile birlestirilebilir
        (17, 'Highlights'),
        (18, 'Application'),
        (19, 'Temperature'),
        (20, 'Construction'),
        (21, 'Norm'),
        (22, 'Available upon request'),
    ]

    desc_val = clean(row_vals[14]) if len(row_vals) > 14 else ''
    wp_val   = clean(row_vals[15]) if len(row_vals) > 15 else ''
    size_val = clean(row_vals[16]) if len(row_vals) > 16 else ''

    parts = []

    # Description of Goods
    if desc_val or wp_val or size_val:
        parts.append('<h4><span style="color: #c0101c;">Description of Goods</span></h4>')
        if desc_val:
            for line in desc_val.split('\n'):
                if line.strip():
                    parts.append(f'<p>{line.strip()}</p>')
        if wp_val:
            wp_lines = [l.strip() for l in wp_val.split('\n') if l.strip()]
            parts.append('<p>WP [Bar] : ' + ', '.join(wp_lines) + '</p>')
        if size_val:
            size_lines = [l.strip() for l in size_val.split('\n') if l.strip()]
            parts.append('<p>Size [mm] [Inch] : ' + ' - '.join(size_lines) + '</p>')

    # Highlights (R=17)
    highlights = clean(row_vals[17]) if len(row_vals) > 17 else ''
    if highlights:
        parts.append('<h4><span style="color: #c0101c;">Highlights</span></h4>')
        lines = [l.strip() for l in highlights.split('\n') if l.strip()]
        parts.append('<p>' + '<br>'.join(lines) + '</p>')

    # Application (S=18)
    application = clean(row_vals[18]) if len(row_vals) > 18 else ''
    if application:
        parts.append('<h4><span style="color: #c0101c;">Application</span></h4>')
        lines = [l.strip() for l in application.split('\n') if l.strip()]
        parts.append('<p>' + ' '.join(lines) + '</p>')

    # Temperature (T=19)
    temperature = clean(row_vals[19]) if len(row_vals) > 19 else ''
    if temperature:
        parts.append('<h4><span style="color: #c0101c;">Temperature</span></h4>')
        lines = [l.strip() for l in temperature.split('\n') if l.strip()]
        parts.append('<p>' + ' '.join(lines) + '</p>')

    # Construction (U=20)
    construction = clean(row_vals[20]) if len(row_vals) > 20 else ''
    if construction:
        parts.append('<h4><span style="color: #c0101c;">Construction</span></h4>')
        lines = [l.strip() for l in construction.split('\n') if l.strip()]
        html_lines = []
        for line in lines:
            # Bold keywords like "Tube:", "Reinforcement:", "Cover:"
            line = re.sub(r'^(Tube|Reinforcement|Cover|Bore|Inner|Outer|Working pressure|Burst pressure|Bend radius)(\s*:)', r'<strong>\1\2</strong>', line)
            html_lines.append(line)
        parts.append('<div class="substrate">' + '<br>'.join(html_lines) + '</div>')

    # Norm (V=21)
    norm = clean(row_vals[21]) if len(row_vals) > 21 else ''
    if norm:
        parts.append('<h4><span style="color: #c0101c;">Norm</span></h4>')
        lines = [l.strip() for l in norm.split('\n') if l.strip()]
        parts.append('<p>' + '<br>'.join(lines) + '</p>')

    # Available upon request (W=22)
    available = clean(row_vals[22]) if len(row_vals) > 22 else ''
    if available:
        parts.append('<h4><span style="color: #c0101c;">Available upon request</span></h4>')
        lines = [l.strip() for l in available.split('\n') if l.strip()]
        parts.append('<p>' + '<br>'.join(lines) + '</p>')

    return '\n'.join(parts)

# --- Excel yukle ---
print("Excel yukleniyor...")
with open(XLSX_PATH, 'rb') as f:
    data = io.BytesIO(f.read())
wb = openpyxl.load_workbook(data, data_only=True)
ws = wb.worksheets[0]
print(f"Satirlar: {ws.max_row}, Sutunlar: {ws.max_column}")

# --- Gorselleri row/col eslesme haritasina don ---
print("Gorseller haritaya aliniyor...")
images = list(ws._images)
# row -> {col -> [image_bytes_list]}
row_col_images = {}
for img in images:
    anchor = img.anchor
    try:
        row = anchor._from.row + 1
        col = anchor._from.col + 1
    except AttributeError:
        continue
    img_bytes = None
    try:
        img_bytes = img.ref.getvalue() if hasattr(img.ref, 'getvalue') else bytes(img.ref)
    except:
        try:
            img_bytes = img._data()
        except:
            pass
    if row not in row_col_images:
        row_col_images[row] = {}
    if col not in row_col_images[row]:
        row_col_images[row][col] = []
    if img_bytes:
        row_col_images[row][col].append(img_bytes)

print(f"Gorsel satirlari: {len(row_col_images)}")

# --- Urunleri isle ---
print("Urunler isleniyor...")
products = []
skipped = 0

for row_num in range(6, ws.max_row + 1):
    row_vals = [ws.cell(row_num, col).value for col in range(1, 25)]

    k = clean(row_vals[10])   # K - category
    m = clean(row_vals[12])   # M - code
    n = clean(row_vals[13])   # N - name

    # Bos satir veya baslik satiri atla
    if not m and not n:
        skipped += 1
        continue
    # M ve N bos, sadece K var (baslik)
    if not m and not n and k:
        skipped += 1
        continue
    # K deger baslik keyword -> ama M veya N varsa urun
    # Devam et - urun satiri

    # Ana gorsel (col B=2)
    main_image_bytes = None
    main_image_ext = 'jpg'
    if row_num in row_col_images and 2 in row_col_images[row_num]:
        main_image_bytes = row_col_images[row_num][2][0]

    # Ikon gorselleri (col X=24)
    icon_images_bytes = []
    if row_num in row_col_images and 24 in row_col_images[row_num]:
        icon_images_bytes = row_col_images[row_num][24]

    # HTML aciklama
    html_desc = build_html_description(row_vals)

    product = {
        'row': row_num,
        'category_raw': k,
        'code': m,
        'name': n,
        'html_description': html_desc,
        'has_main_image': main_image_bytes is not None,
        'icon_count': len(icon_images_bytes),
    }
    products.append(product)

    # Ana gorseli kaydet
    if main_image_bytes:
        img_fname = f'row{row_num}_B.jpg'
        img_path = os.path.join(IMG_DIR, img_fname)
        with open(img_path, 'wb') as f:
            f.write(main_image_bytes)
        product['main_image_file'] = img_fname

    # Ikon gorsellerini kaydet
    icon_files = []
    for i, ib in enumerate(icon_images_bytes):
        icon_fname = f'row{row_num}_X_{i}.jpg'
        icon_path = os.path.join(IMG_DIR, icon_fname)
        with open(icon_path, 'wb') as f:
            f.write(ib)
        icon_files.append(icon_fname)
    product['icon_files'] = icon_files

print(f"Toplam urun: {len(products)}, Atlanan: {skipped}")

# Unique category values
cat_vals = {}
for p in products:
    cr = p['category_raw'].upper().strip()
    if cr not in cat_vals:
        cat_vals[cr] = 0
    cat_vals[cr] += 1
print(f"\nKategori dagilimi:")
for cat, cnt in sorted(cat_vals.items()):
    print(f"  '{cat}': {cnt} urun")

# JSON kaydet
out_json = os.path.join(OUT_DIR, 'products.json')
with open(out_json, 'w', encoding='utf-8') as f:
    json.dump(products, f, indent=2, ensure_ascii=False)
print(f"\nUrunler kaydedildi: {out_json}")
print(f"Gorsel dizini: {IMG_DIR}")
