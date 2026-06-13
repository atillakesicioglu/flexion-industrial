# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import json, io, os

XLSX_PATH = r'C:\Users\atill\OneDrive\MASAST~1\projects\flexion\RNLER~1.XLS'

with open(XLSX_PATH, 'rb') as f:
    data = io.BytesIO(f.read())
wb = openpyxl.load_workbook(data, data_only=True)
ws = wb.worksheets[0]

print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

# Row 2 headers
print("\n=== Row 2 Headers ===")
for col in range(1, ws.max_column + 1):
    v = ws.cell(2, col).value
    if v is not None:
        print(f"  {get_column_letter(col)}({col}): {v}")

# Sample rows 3-6 to understand structure
print("\n=== Sample rows 3-6 (cols K-X) ===")
for row_num in range(3, 7):
    print(f"\n  Row {row_num}:")
    for col in range(11, 25):  # K=11 to X=24
        v = ws.cell(row_num, col).value
        if v is not None:
            lbl = get_column_letter(col)
            val_str = str(v)[:80]
            print(f"    {lbl}({col}): {val_str}")

# Image anchor analysis
print("\n=== Image anchor analysis (first 20) ===")
images = list(ws._images)
print(f"Total images: {len(images)}")

# Map: row -> list of images
row_to_images = {}
for img in images:
    anchor = img.anchor
    try:
        # TwoCellAnchor
        row = anchor._from.row + 1  # 0-indexed to 1-indexed
        col = anchor._from.col + 1
    except AttributeError:
        try:
            # OneCellAnchor
            row = anchor._from.row + 1
            col = anchor._from.col + 1
        except:
            row = -1
            col = -1
    if row not in row_to_images:
        row_to_images[row] = []
    row_to_images[row].append({'col': col, 'col_letter': get_column_letter(col) if col > 0 else '?'})

print(f"Rows with images: {len(row_to_images)}")
print("First 10 row->image mappings:")
for r in sorted(row_to_images.keys())[:10]:
    print(f"  Row {r}: {row_to_images[r]}")

# Check how many rows have col B (2) images vs col X (24) images
col_b_imgs = sum(1 for r, imgs in row_to_images.items() for im in imgs if im['col'] == 2)
col_x_imgs = sum(1 for r, imgs in row_to_images.items() for im in imgs if im['col'] == 24)
print(f"\nImages in col B(2): {col_b_imgs}")
print(f"Images in col X(24): {col_x_imgs}")

# Distinct column positions of images
all_cols = set()
for r, imgs in row_to_images.items():
    for im in imgs:
        all_cols.add(im['col'])
print(f"Distinct image columns: {sorted(all_cols)}")

# Save row_to_images mapping as JSON
out_path = r'C:\Users\atill\OneDrive\MASAST~1\projects\flexion\scripts\row_image_map.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({str(k): v for k, v in row_to_images.items()}, f, indent=2, ensure_ascii=False)
print(f"\nRow-image map saved to: {out_path}")
