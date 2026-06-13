# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import json, os

XLSX_PATH = r'C:\Users\atill\OneDrive\MASAST~1\projects\flexion\RNLER~1.XLS'

import io
with open(XLSX_PATH, 'rb') as f:
    data = io.BytesIO(f.read())
wb = openpyxl.load_workbook(data, data_only=True)
print("Sheet names:", wb.sheetnames)
ws = wb.worksheets[0]
print("Active sheet:", ws.title)
print("Max row:", ws.max_row, "Max col:", ws.max_column)

# Print row 2 (headers)
print("\n--- Row 2 (headers) ---")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    if cell.value is not None:
        print(f"  Col {get_column_letter(col)} ({col}): {cell.value}")

# Print first data row (row 3)
print("\n--- Row 3 (first data row) ---")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(3, col)
    if cell.value is not None:
        val = str(cell.value)[:120]
        print(f"  Col {get_column_letter(col)} ({col}): {val}")

# Count non-empty rows from row 3
count = 0
for row in ws.iter_rows(min_row=3):
    if any(c.value for c in row):
        count += 1
print(f"\nTotal data rows (row 3+): {count}")

# Category values (col K = index 11 in 1-based)
print("\n--- Category values (col K=11) ---")
cats = set()
for row_vals in ws.iter_rows(min_row=3, values_only=True):
    k_val = row_vals[10] if len(row_vals) > 10 else None
    if k_val:
        cats.add(str(k_val).strip())
for c in sorted(cats):
    print(f"  '{c}'")

# Check embedded images
images = list(ws._images)
print(f"\n--- Embedded images: {len(images)} ---")
for i, img in enumerate(images[:5]):
    anchor = img.anchor
    print(f"  Image[{i}]: anchor={anchor}, type={type(img).__name__}")
